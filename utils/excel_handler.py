import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

from config.config import EXCEL_FILE

class ExcelHandler:
    def __init__(self):
        self.filename = EXCEL_FILE
        self.workbook = None
        
        try:
            # Check if the file exists and load or create a new one
            if os.path.exists(self.filename):
                self.workbook = load_workbook(self.filename)
            else:
                self.workbook = Workbook()  # Create new workbook if file doesn't exist

        except Exception as e:
            print(f"Error initializing Excel file: {e}")
            raise

    def add_result(self, url, test_name, status, comments):
        try:
            # Check if the sheet exists, if not create it
            if test_name not in self.workbook.sheetnames:
                sheet = self.workbook.create_sheet(test_name)
                # Add headers
                headers = ['Timestamp', 'URL', 'Test Case', 'Status', 'Comments']
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=col, value=header)
            else:
                sheet = self.workbook[test_name]

            # Get the next empty row
            next_row = sheet.max_row + 1

            # Prepare the data
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data = [timestamp, url, test_name, status, comments]

            # Add data to the next row
            for col, value in enumerate(data, 1):
                sheet.cell(row=next_row, column=col, value=value)

            # Adjust column widths after adding the new row
            self._adjust_column_width(sheet)

            # Save the workbook
            self.workbook.save(self.filename)
            
        except Exception as e:
            print(f"Error adding result to Excel: {e}")
            # If there's an error, create a new workbook and try adding the result again
            self.workbook = Workbook()
            self.add_result(url, test_name, status, comments)

    def _adjust_column_width(self, sheet):
        """ Adjust column widths based on the longest data in each column. """
        for column in sheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def close(self):
        try:
            # Remove the default "Sheet" if it exists (usually an empty sheet)
            if "Sheet" in self.workbook.sheetnames:
                self.workbook.remove(self.workbook["Sheet"])

            # Save the workbook after removing the initial sheet
            if self.workbook:
                self.workbook.save(self.filename)
                
        except Exception as e:
            print(f"Error closing Excel file: {e}")
